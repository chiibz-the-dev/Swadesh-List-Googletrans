from googletrans import Translator
import openpyxl

# Swadesh list in an array
swadesh_list = [
    "I", "you (singular)", "he", "we", "you (plural)", "they", "this", "that", "here", "there",
    "who", "what", "where", "when", "how", "not", "all", "many", "some", "few", "other", "one", "two",
    "three", "four", "five", "big", "long", "wide", "thick", "heavy", "small", "short", "narrow", "thin",
    "woman", "man (adult male)", "man (human being)", "child", "wife", "husband", "mother", "father", "animal",
    "fish", "bird", "dog", "louse", "snake", "worm", "tree", "forest", "stick", "fruit", "seed", "leaf", "root",
    "bark (of a tree)", "flower", "grass", "rope", "skin", "meat", "blood", "bone", "fat (noun)", "egg", "horn",
    "tail", "feather", "hair", "head", "ear", "eye", "nose", "mouth", "tooth", "tongue (organ)", "fingernail",
    "foot", "leg", "knee", "hand", "wing", "belly", "guts", "neck", "back", "breast", "heart", "liver", "to drink",
    "to eat", "to bite", "to suck", "to spit", "to vomit", "to blow", "to breathe", "to laugh", "to see", "to hear",
    "to know", "to think", "to smell", "to fear", "to sleep", "to live", "to die", "to kill", "to fight", "to hunt",
    "to hit", "to cut", "to split", "to stab", "to scratch", "to dig", "to swim", "to fly", "to walk", "to come",
    "to lie (as in a bed)", "to sit", "to stand", "to turn (intransitive)", "to fall", "to give", "to hold", "to squeeze",
    "to rub", "to wash", "to wipe", "to pull", "to push", "to throw", "to tie", "to sew", "to count", "to say", "to sing",
    "to play", "to float", "to flow", "to freeze", "to swell", "sun", "moon", "star", "water", "rain", "river", "lake",
    "sea", "salt", "stone", "sand", "dust", "earth", "cloud", "fog", "sky", "wind", "snow", "ice", "smoke", "fire",
    "ash", "to burn", "road", "mountain", "red", "green", "yellow", "white", "black", "night", "day", "year", "warm",
    "cold", "full", "new", "old", "good", "bad", "rotten", "dirty", "straight", "round", "sharp (as a knife)",
    "dull (as a knife)", "smooth", "wet", "dry", "correct", "near", "far", "right", "left", "at", "in", "with", "and",
    "if", "because", "name"
]

# Create a new workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Swadesh List - Deutsch"
#change language name

# Write the Swadesh list in the first column
for i, word in enumerate(swadesh_list, 1):
    ws.cell(row=i, column=1, value=word)

# Initialize Google Translator
translator = Translator()

# Translate each word into German and write to a new column in the Excel file
# Switch from 'de' to target language
for i, word in enumerate(swadesh_list, 1):
    translation = translator.translate(word, src='en', dest='de').text
    ws.cell(row=i, column=2, value=translation)

# Save the workbook, name purposefully
file_path = 'Swadesh_List_Deutsch.xlsx'
wb.save(file_path)
print(f"File saved at: {file_path}")
